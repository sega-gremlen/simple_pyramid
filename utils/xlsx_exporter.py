import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_excel_from_json(json_data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"

    # 1. Собираем все даты
    all_dates = set()
    for point in json_data:
        for param in point.get("parametersData", []):
            for val in param.get("values", []):
                dt_str = val.get("valueDt")
                if dt_str:
                    dt = datetime.fromisoformat(dt_str.replace('Z', ''))
                    all_dates.add(dt)

    sorted_dates = sorted(all_dates)

    # 2. Заголовки
    headers = ["Прибор учета", "Параметр"]
    headers.extend([d.strftime("%d.%m.%Y 00:00") for d in sorted_dates])
    ws.append(headers)

    # 3. Форматирование первой строки (заголовки): жирный, по центру
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_idx = 2  # следующая строка для "Точка учета"

    for point in json_data:
        point_caption = point["pointWithMeter"]["caption"]
        meter_caption = point["pointWithMeter"]["meter"]["caption"]

        # Строка с точкой учёта (вторая строка)
        cell = ws.cell(row=row_idx, column=1, value=f"Точка учета: {point_caption}")
        cell.font = Font(bold=True)   # жирный шрифт, центровка оставлена по умолчанию (лево)
        row_idx += 1

        # Параметры
        for param in point.get("parametersData", []):
            # Формируем строку для столбца B: "caption, unit"
            param_caption = param["parameter"]["caption"]
            param_unit = "кВт*ч"
            param_display = f"{param_caption}, {param_unit}"

            value_by_date = {}
            for val in param.get("values", []):
                dt_str = val.get("valueDt")
                if dt_str:
                    dt = datetime.fromisoformat(dt_str.replace('Z', ''))
                    raw_value = val.get("value")
                    if raw_value is not None:
                        value_by_date[dt] = raw_value / 1000.0

            row_data = [meter_caption, param_display]
            for d in sorted_dates:
                v = value_by_date.get(d)
                row_data.append(v if v is not None else "")
            ws.append(row_data)
            row_idx += 1

    # 4. Ширина всех столбцов = 20
    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    # 5. Перенос текста для столбцов A и B во всех строках, начиная с 3-й
    for row in range(3, ws.max_row + 1):
        for col in [1, 2]:  # A = 1, B = 2
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
    # 6. Устанавливаем высоту строк 3, 4, 5 = 60
    for row in [3, 4, 5]:
        if row <= ws.max_row:  # проверяем, существует ли строка
            ws.row_dimensions[row].height = 60
    
    wb.save(output_path)


if __name__ == "__main__":
    with open("input.json", 'r', encoding='utf-8') as f:
        data = json.load(f)
        
        path = "C:/Users/FilippovAS/Desktop"
        title = "output.xlsx"
        
    output_path = paht / title

    create_excel_from_json(data, output_path)
    print(f"Excel file saved to {output_path}")